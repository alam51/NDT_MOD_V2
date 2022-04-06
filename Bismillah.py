import dateutil.parser as date_parser
import os
import openpyxl as xl
from openpyxl import Workbook
import time
import multiprocessing


def DOD_lister(wb_list, new_wb_full_path, new_wb_index):
    wb_list[new_wb_index] = xl.load_workbook(new_wb_full_path, data_only=True)


if __name__ == '__main__':
    start = time.perf_counter()

    data_month = 3
    path = r'F:\IMD\MOD\March 2021_Final\March 2021_Final\March2021'
    files = os.listdir(path)
    day = []
    month = []
    count = 1
    # w = xl.load_workbook('A.xlsx', data_only=True)
    w = Workbook()
    wb_list = [w]
    for i in range(1, 32 + 1):
        wb_list.append(w)
    print(len(wb_list))

    processes = []
    for file in files:
        try:
            d = date_parser.parse(file, fuzzy=True, dayfirst=True)
            day.append(d.day)
            month.append(d.month)

            if d.month == data_month:
                # wb_list[d.day] = xl.load_workbook(path + '\\' + file, data_only=True)
                p = multiprocessing.Process(target=DOD_lister, args=(wb_list, path + '\\' + file, d.day))
                p.start()
                processes.append(p)
            # elif d.month == data_month + 1 and d.day == 1:
            #     wb_list[32] = xl.load_workbook(path + '\\' + file, data_only=True)


        except ValueError:
            print('No valid date found at file name: ' + file)

    for process in processes:
        process.join()

    end = time.perf_counter()

    print(day)
    print(month)
    print(len(month))
    print(wb_list[5]['EWIC']['D35'].value)
    print(wb_list[10]['P1']['H4'].value)
    print(wb_list[15]['Forecast']['E3'].value)
    print(f'Time taken {end-start}')
